from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Prefetch
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from academico.models import (
    Asignatura,
    Aula,
    CentroTutorial,
    Facultad,
    Modalidad,
    PeriodoAcademico,
    PlanEstudioAsignatura,
    Profesor,
    ProgramaAcademico,
    ProgramaCentroTutorial,
    Sede,
    Semestre,
)

from .forms import HorarioForm, NuevaAsignacionForm
from .models import (
    Grupo,
    Horario,
    OfertaAcademica,
    OfertaGrupo,
)

# ==========================================================
# CONFIGURACIÓN DEL CALENDARIO
# ==========================================================

HORA_INICIO_CALENDARIO = 6
HORA_FIN_CALENDARIO = 22
PIXELES_POR_MINUTO = 0.75


# ==========================================================
# FUNCIONES AUXILIARES
# ==========================================================


def agregar_errores_formulario(form, error):
    """
    Agrega los errores de validación del modelo como
    errores generales del formulario.

    De esta forma los conflictos de profesor, aula,
    grupo, etc. se muestran con el mismo estilo visual
    en la parte superior del formulario.
    """

    if hasattr(error, "message_dict"):

        for errores in error.message_dict.values():

            for mensaje in errores:

                form.add_error(
                    None,
                    mensaje,
                )

    else:

        mensajes = getattr(
            error,
            "messages",
            [str(error)],
        )

        for mensaje in mensajes:

            form.add_error(
                None,
                mensaje,
            )


# ==========================================================
# PANTALLA DE INICIO
# ==========================================================


@login_required
def inicio(request):

    context = {
        "centros": CentroTutorial.objects.filter(activo=True).order_by("nombre"),
        "total_programas": ProgramaAcademico.objects.filter(activo=True).count(),
        "total_profesores": Profesor.objects.filter(activo=True).count(),
        "total_aulas": Aula.objects.filter(activo=True).count(),
        "total_horarios": Horario.objects.filter(activo=True).count(),
    }

    return render(
        request,
        "horarios/inicio.html",
        context,
    )


# ==========================================================
# PLANIFICACIÓN / CALENDARIO SEMANAL
# ==========================================================


@login_required
@permission_required(
    "horarios.view_horario",
    raise_exception=True,
)
def planificacion(request):

    # ======================================================
    # FILTROS
    # ======================================================

    centro_id = request.GET.get("centro", "").strip()
    sede_id = request.GET.get("sede", "").strip()
    facultad_id = request.GET.get("facultad", "").strip()
    programa_id = request.GET.get("programa", "").strip()
    periodo_id = request.GET.get("periodo", "").strip()
    semestre_id = request.GET.get("semestre", "").strip()
    grupo_id = request.GET.get("grupo", "").strip()

    # ======================================================
    # FILTROS OBLIGATORIOS
    #
    # Sede y Grupo son opcionales.
    # ======================================================

    filtros_completos = all(
        [
            centro_id,
            facultad_id,
            programa_id,
            periodo_id,
            semestre_id,
        ]
    )

    # ======================================================
    # OBJETOS SELECCIONADOS
    # ======================================================

    centro = None
    sede = None
    facultad = None
    programa = None
    periodo = None
    semestre = None
    grupo = None

    horarios = []

    # ======================================================
    # CARGAR PLANIFICACIÓN
    # ======================================================

    if filtros_completos:

        centro = get_object_or_404(
            CentroTutorial,
            pk=centro_id,
            activo=True,
        )

        facultad = get_object_or_404(
            Facultad,
            pk=facultad_id,
            activo=True,
        )

        programa = get_object_or_404(
            ProgramaAcademico,
            pk=programa_id,
            facultad=facultad,
            activo=True,
        )

        # Verificamos que el programa esté disponible
        # en el Centro Tutorial seleccionado.

        get_object_or_404(
            ProgramaCentroTutorial,
            programa=programa,
            centro_tutorial=centro,
            activo=True,
        )

        periodo = get_object_or_404(
            PeriodoAcademico,
            pk=periodo_id,
            activo=True,
        )

        semestre = get_object_or_404(
            Semestre,
            pk=semestre_id,
        )

        # ==================================================
        # SEDE OPCIONAL
        # ==================================================

        if sede_id:

            sede = get_object_or_404(
                Sede,
                pk=sede_id,
                centro_tutorial=centro,
                activo=True,
            )

        # ==================================================
        # GRUPO OPCIONAL
        # ==================================================

        if grupo_id:

            grupo = get_object_or_404(
                Grupo,
                pk=grupo_id,
                centro_tutorial=centro,
                programa=programa,
                periodo=periodo,
                semestre=semestre,
                activo=True,
            )

        # ==================================================
        # CONSULTA BASE
        # ==================================================

        horarios_query = Horario.objects.filter(
            activo=True,
            oferta__activo=True,
            oferta__periodo=periodo,
            oferta__grupos__centro_tutorial=centro,
            oferta__grupos__programa=programa,
            oferta__grupos__semestre=semestre,
        )

        # ==================================================
        # FILTRO DE SEDE
        #
        # Solamente se aplica si el usuario seleccionó una.
        # ==================================================

        if sede:

            horarios_query = horarios_query.filter(
                aula__sede=sede,
            )

        # ==================================================
        # FILTRO DE GRUPO
        # ==================================================

        if grupo:

            horarios_query = horarios_query.filter(
                oferta__grupos=grupo,
            )

        horarios = list(
            horarios_query.select_related(
                "oferta",
                "oferta__asignatura",
                "oferta__profesor",
                "oferta__periodo",
                "oferta__modalidad",
                "aula",
                "aula__sede",
            )
            .prefetch_related(
                "oferta__grupos",
            )
            .distinct()
            .order_by(
                "dia",
                "hora_inicio",
            )
        )

    # ======================================================
    # DÍAS
    # ======================================================

    dias = [
        ("LU", "Lunes"),
        ("MA", "Martes"),
        ("MI", "Miércoles"),
        ("JU", "Jueves"),
        ("VI", "Viernes"),
        ("SA", "Sábado"),
        ("DO", "Domingo"),
    ]

    # ======================================================
    # DIMENSIONES DEL CALENDARIO
    # ======================================================

    inicio_calendario = HORA_INICIO_CALENDARIO * 60
    fin_calendario = HORA_FIN_CALENDARIO * 60

    altura_calendario = int((fin_calendario - inicio_calendario) * PIXELES_POR_MINUTO)

    # ======================================================
    # HORAS DEL CALENDARIO
    # ======================================================

    horas_calendario = []

    for hora in range(
        HORA_INICIO_CALENDARIO,
        HORA_FIN_CALENDARIO,
    ):

        minutos = hora * 60

        top = int((minutos - inicio_calendario) * PIXELES_POR_MINUTO)

        horas_calendario.append(
            {
                "texto": f"{hora:02d}:00",
                "top": top,
            }
        )

    # ======================================================
    # CONSTRUIR DÍAS DEL CALENDARIO
    # ======================================================

    dias_calendario = []

    for codigo, nombre in dias:

        bloques = []

        for horario in horarios:

            if horario.dia != codigo:
                continue

            inicio = horario.hora_inicio.hour * 60 + horario.hora_inicio.minute

            fin = horario.hora_fin.hour * 60 + horario.hora_fin.minute

            if fin <= inicio_calendario or inicio >= fin_calendario:
                continue

            inicio_visible = max(
                inicio,
                inicio_calendario,
            )

            fin_visible = min(
                fin,
                fin_calendario,
            )

            top = int((inicio_visible - inicio_calendario) * PIXELES_POR_MINUTO)

            altura = max(
                int((fin_visible - inicio_visible) * PIXELES_POR_MINUTO),
                30,
            )

            # ==================================================
            # INFORMACIÓN DE GRUPOS
            # ==================================================

            grupos_oferta = list(horario.oferta.grupos.all())

            es_transversal = len(grupos_oferta) > 1

            # ==================================================
            # COLOR DEL EVENTO
            # ==================================================

            if es_transversal:

                clase_evento = "evento-transversal"

            else:

                codigo_modalidad = horario.oferta.modalidad.codigo.strip().upper()

                if codigo_modalidad == "VIR":

                    clase_evento = "evento-virtual"

                elif codigo_modalidad == "HIB":

                    clase_evento = "evento-hibrida"

                else:

                    clase_evento = "evento-presencial"

            bloques.append(
                {
                    "horario": horario,
                    "top": top,
                    "altura": altura,
                    "clase_evento": clase_evento,
                    "es_transversal": es_transversal,
                    "grupos": grupos_oferta,
                }
            )

        dias_calendario.append(
            {
                "codigo": codigo,
                "nombre": nombre,
                "horarios": bloques,
            }
        )

    # ======================================================
    # CONTEXTO
    # ======================================================

    context = {
        # Datos iniciales
        "centros": (CentroTutorial.objects.filter(activo=True).order_by("nombre")),
        # Objetos seleccionados
        "centro": centro,
        "sede": sede,
        "facultad": facultad,
        "programa": programa,
        "periodo": periodo,
        "semestre": semestre,
        "grupo": grupo,
        # IDs seleccionados
        "centro_id": centro_id,
        "sede_id": sede_id,
        "facultad_id": facultad_id,
        "programa_id": programa_id,
        "periodo_id": periodo_id,
        "semestre_id": semestre_id,
        "grupo_id": grupo_id,
        # Estado
        "filtros_completos": filtros_completos,
        "hay_horarios": bool(horarios),
        # Calendario
        "dias_calendario": dias_calendario,
        "horas_calendario": horas_calendario,
        "altura_calendario": altura_calendario,
    }

    return render(
        request,
        "horarios/planificacion.html",
        context,
    )


# ==========================================================
# SELECTS DEPENDIENTES - AJAX
# ==========================================================


@login_required
def cargar_sedes(request):
    centro_id = request.GET.get("centro")

    sedes = Sede.objects.filter(
        centro_tutorial_id=centro_id,
        activo=True,
    ).values(
        "id",
        "nombre",
    )

    return JsonResponse(
        list(sedes),
        safe=False,
    )


@login_required
def cargar_aulas(request):

    sede_id = request.GET.get("sede")

    if not sede_id:

        return JsonResponse(
            [],
            safe=False,
        )

    aulas = (
        Aula.objects.filter(
            sede_id=sede_id,
            activo=True,
        )
        .order_by(
            "nombre",
        )
        .values(
            "id",
            "nombre",
            "capacidad",
            "tipo",
        )
    )

    return JsonResponse(
        list(aulas),
        safe=False,
    )


@login_required
def cargar_facultades(request):
    centro_id = request.GET.get("centro")
    sede_id = request.GET.get("sede")

    if not centro_id and sede_id:
        centro_id = (
            Sede.objects.filter(pk=sede_id)
            .values_list("centro_tutorial_id", flat=True)
            .first()
        )

    if not centro_id:
        return JsonResponse([], safe=False)

    facultades = (
        Facultad.objects.filter(
            activo=True,
            programas_academicos__activo=True,
            programas_academicos__centros_tutoriales__centro_tutorial_id=centro_id,
            programas_academicos__centros_tutoriales__activo=True,
        )
        .distinct()
        .values(
            "id",
            "nombre",
        )
    )

    return JsonResponse(
        list(facultades),
        safe=False,
    )


@login_required
def cargar_programas(request):
    facultad_id = request.GET.get("facultad")
    centro_id = request.GET.get("centro")
    sede_id = request.GET.get("sede")

    if not centro_id and sede_id:
        centro_id = (
            Sede.objects.filter(pk=sede_id)
            .values_list("centro_tutorial_id", flat=True)
            .first()
        )

    programas = ProgramaAcademico.objects.filter(
        facultad_id=facultad_id,
        activo=True,
    )

    if centro_id:
        programas = programas.filter(
            centros_tutoriales__centro_tutorial_id=centro_id,
            centros_tutoriales__activo=True,
        )

    programas = programas.distinct().values(
        "id",
        "nombre",
    )

    return JsonResponse(
        list(programas),
        safe=False,
    )


@login_required
def cargar_periodos(request):

    centro_id = request.GET.get("centro")
    programa_id = request.GET.get("programa")

    periodos = PeriodoAcademico.objects.filter(
        activo=True,
    )

    if centro_id and programa_id:

        periodos = periodos.filter(
            grupos__centro_tutorial_id=centro_id,
            grupos__programa_id=programa_id,
            grupos__activo=True,
        )

    periodos = (
        periodos.distinct()
        .order_by("-codigo")
        .values(
            "id",
            "nombre",
        )
    )

    return JsonResponse(
        list(periodos),
        safe=False,
    )


@login_required
def cargar_semestres(request):

    centro_id = request.GET.get("centro")
    programa_id = request.GET.get("programa")
    periodo_id = request.GET.get("periodo")

    semestres = Semestre.objects.all()

    if centro_id and programa_id and periodo_id:

        semestres = semestres.filter(
            grupos__centro_tutorial_id=centro_id,
            grupos__programa_id=programa_id,
            grupos__periodo_id=periodo_id,
            grupos__activo=True,
        )

    elif programa_id:

        semestres = semestres.filter(
            planes_estudio__programa_id=programa_id,
            planes_estudio__activo=True,
        )

    semestres = semestres.distinct().order_by("numero")

    datos = [
        {
            "id": semestre.id,
            "numero": semestre.numero,
            "nombre": semestre.get_numero_display(),
        }
        for semestre in semestres
    ]

    return JsonResponse(
        datos,
        safe=False,
    )


@login_required
def cargar_grupos(request):
    centro_id = request.GET.get("centro")
    programa_id = request.GET.get("programa") or request.GET.get("carrera")
    periodo_id = request.GET.get("periodo")
    semestre_id = request.GET.get("semestre")

    grupos = Grupo.objects.filter(activo=True)

    if centro_id:
        grupos = grupos.filter(centro_tutorial_id=centro_id)

    if programa_id:
        grupos = grupos.filter(programa_id=programa_id)

    if periodo_id:
        grupos = grupos.filter(periodo_id=periodo_id)

    if semestre_id:
        grupos = grupos.filter(semestre_id=semestre_id)

    grupos = grupos.values(
        "id",
        "codigo",
    )

    return JsonResponse(
        list(grupos),
        safe=False,
    )


# ==========================================================
# NUEVA ASIGNACIÓN
# ==========================================================


@login_required
@permission_required(
    "horarios.add_horario",
    raise_exception=True,
)
def nueva_asignacion(request):

    # ======================================================
    # CONTEXTO ACADÉMICO
    # ======================================================

    centro_id = request.GET.get("centro") or request.POST.get("centro")

    periodo_id = request.GET.get("periodo") or request.POST.get("periodo")

    semestre_id = request.GET.get("semestre") or request.POST.get("semestre")

    programa_id = request.GET.get("programa") or request.POST.get("programa")

    sede_id = request.GET.get("sede") or request.POST.get("sede")

    # ------------------------------------------------------
    # Centro, Programa, Periodo y Semestre son necesarios.
    #
    # Sede es opcional porque ahora puede seleccionarse
    # directamente dentro del formulario.
    # ------------------------------------------------------

    if not all(
        [
            centro_id,
            programa_id,
            periodo_id,
            semestre_id,
        ]
    ):

        messages.error(
            request,
            ("Falta información académica para " "crear la asignación."),
        )

        return redirect("planificacion")

    # ======================================================
    # OBTENER CONTEXTO
    # ======================================================

    centro = get_object_or_404(
        CentroTutorial,
        pk=centro_id,
        activo=True,
    )

    programa = get_object_or_404(
        ProgramaAcademico,
        pk=programa_id,
        activo=True,
    )

    periodo = get_object_or_404(
        PeriodoAcademico,
        pk=periodo_id,
        activo=True,
    )

    semestre = get_object_or_404(
        Semestre,
        pk=semestre_id,
    )

    # ------------------------------------------------------
    # Verificar que el programa esté disponible
    # en el Centro Tutorial.
    # ------------------------------------------------------

    get_object_or_404(
        ProgramaCentroTutorial,
        programa=programa,
        centro_tutorial=centro,
        activo=True,
    )

    # ======================================================
    # SEDE INICIAL OPCIONAL
    # ======================================================

    sede_inicial = None

    if sede_id:

        sede_inicial = get_object_or_404(
            Sede,
            pk=sede_id,
            centro_tutorial=centro,
            activo=True,
        )

    # ======================================================
    # URL DE RETORNO
    # ======================================================

    volver_a = (
        request.POST.get("volver_a")
        or request.GET.get("volver_a")
        or request.META.get("HTTP_REFERER")
        or "/"
    )

    # ======================================================
    # POST
    # ======================================================

    if request.method == "POST":

        form = NuevaAsignacionForm(
            request.POST,
            programa=programa,
            semestre=semestre,
            centro_tutorial=centro,
            periodo=periodo,
            sede_inicial=sede_inicial,
        )

        if form.is_valid():

            asignatura = form.cleaned_data["asignatura"]

            profesor = form.cleaned_data["profesor"]

            modalidad = form.cleaned_data["modalidad"]

            grupos = form.cleaned_data["grupos"]

            cupos = form.cleaned_data["cupos"]

            aula = form.cleaned_data["aula"]

            dia = form.cleaned_data["dia"]

            hora_inicio = form.cleaned_data["hora_inicio"]

            hora_fin = form.cleaned_data["hora_fin"]

            try:

                # ==========================================
                # TRANSACCIÓN
                #
                # Si falla cualquier validación,
                # no queda información parcial guardada.
                # ==========================================

                with transaction.atomic():

                    # ======================================
                    # 1. CREAR OFERTA ACADÉMICA
                    # ======================================

                    oferta = OfertaAcademica(
                        asignatura=asignatura,
                        profesor=profesor,
                        periodo=periodo,
                        modalidad=modalidad,
                        activo=True,
                    )

                    oferta.full_clean()
                    oferta.save()

                    # ======================================
                    # 2. ASOCIAR GRUPOS
                    # ======================================

                    for grupo in grupos:

                        oferta_grupo = OfertaGrupo(
                            oferta=oferta,
                            grupo=grupo,
                            cupos=cupos,
                        )

                        oferta_grupo.full_clean()
                        oferta_grupo.save()

                    # ======================================
                    # 3. CREAR HORARIO
                    # ======================================

                    horario = Horario(
                        oferta=oferta,
                        aula=aula,
                        dia=dia,
                        hora_inicio=hora_inicio,
                        hora_fin=hora_fin,
                        activo=True,
                    )

                    # --------------------------------------
                    # Aquí se ejecutan las validaciones de:
                    #
                    # - profesor
                    # - aula
                    # - grupos
                    # - superposición horaria
                    # --------------------------------------

                    horario.full_clean()

                    horario.save()

            except ValidationError as error:

                agregar_errores_formulario(
                    form,
                    error,
                )

            else:

                cantidad_grupos = grupos.count()

                if cantidad_grupos > 1:

                    mensaje = "Asignación transversal creada " "correctamente."

                else:

                    mensaje = "Asignación creada " "correctamente."

                messages.success(
                    request,
                    mensaje,
                )

                return redirect(volver_a)

    # ======================================================
    # GET
    # ======================================================

    else:

        form = NuevaAsignacionForm(
            programa=programa,
            semestre=semestre,
            centro_tutorial=centro,
            periodo=periodo,
            sede_inicial=sede_inicial,
        )

    # ======================================================
    # CONTEXTO
    # ======================================================

    context = {
        "form": form,
        "centro": centro,
        "programa": programa,
        "periodo": periodo,
        "semestre": semestre,
        "sede": sede_inicial,
        "volver_a": volver_a,
    }

    return render(
        request,
        "horarios/nueva_asignacion.html",
        context,
    )


# ==========================================================
# EDITAR ASIGNACIÓN
# ==========================================================


@login_required
@permission_required(
    "horarios.change_horario",
    raise_exception=True,
)
def editar_asignacion(
    request,
    horario_id,
):

    # ======================================================
    # HORARIO ACTUAL
    # ======================================================

    horario = get_object_or_404(
        Horario.objects.select_related(
            "oferta",
            "oferta__asignatura",
            "oferta__profesor",
            "oferta__periodo",
            "oferta__modalidad",
            "aula",
            "aula__sede",
            "aula__sede__centro_tutorial",
        ).prefetch_related(
            "oferta__grupos",
        ),
        pk=horario_id,
    )

    oferta = horario.oferta

    # ======================================================
    # GRUPOS ACTUALES
    # ======================================================

    grupos_actuales = list(
        oferta.grupos.select_related(
            "programa",
            "centro_tutorial",
            "periodo",
            "semestre",
        ).all()
    )

    if not grupos_actuales:

        messages.error(
            request,
            ("La oferta académica no tiene " "grupos asociados."),
        )

        return redirect("planificacion")

    # ------------------------------------------------------
    # Utilizamos el primer grupo como contexto de referencia.
    #
    # En una asignación transversal pueden existir varios
    # programas, pero todos pertenecen al mismo periodo,
    # centro y semestre.
    # ------------------------------------------------------

    grupo_referencia = grupos_actuales[0]

    centro = grupo_referencia.centro_tutorial

    programa = grupo_referencia.programa

    periodo = oferta.periodo

    semestre = grupo_referencia.semestre

    sede_actual = horario.aula.sede

    # ======================================================
    # URL DE RETORNO
    # ======================================================

    volver_a = (
        request.POST.get("volver_a")
        or request.GET.get("volver_a")
        or request.META.get("HTTP_REFERER")
        or "/"
    )

    # ======================================================
    # CUPOS ACTUALES
    # ======================================================

    oferta_grupos_actuales = list(
        OfertaGrupo.objects.filter(
            oferta=oferta,
        ).select_related(
            "grupo",
            "grupo__programa",
        )
    )

    if oferta_grupos_actuales:

        cupos_actuales = oferta_grupos_actuales[0].cupos

    else:

        cupos_actuales = 30

    # ======================================================
    # POST
    # ======================================================

    if request.method == "POST":

        form = NuevaAsignacionForm(
            request.POST,
            programa=programa,
            semestre=semestre,
            centro_tutorial=centro,
            periodo=periodo,
            sede_inicial=sede_actual,
        )

        if form.is_valid():

            asignatura = form.cleaned_data["asignatura"]

            profesor = form.cleaned_data["profesor"]

            modalidad = form.cleaned_data["modalidad"]

            grupos = form.cleaned_data["grupos"]

            cupos = form.cleaned_data["cupos"]

            aula = form.cleaned_data["aula"]

            dia = form.cleaned_data["dia"]

            hora_inicio = form.cleaned_data["hora_inicio"]

            hora_fin = form.cleaned_data["hora_fin"]

            try:

                # ==========================================
                # TRANSACCIÓN
                #
                # Si una validación falla, toda la edición
                # vuelve automáticamente a su estado previo.
                # ==========================================

                with transaction.atomic():

                    # ======================================
                    # 1. ACTUALIZAR OFERTA
                    # ======================================

                    oferta.asignatura = asignatura
                    oferta.profesor = profesor
                    oferta.modalidad = modalidad
                    oferta.periodo = periodo
                    oferta.activo = True

                    oferta.full_clean()

                    oferta.save()

                    # ======================================
                    # 2. ACTUALIZAR GRUPOS
                    #
                    # Eliminamos las asociaciones anteriores
                    # y reconstruimos las seleccionadas.
                    # ======================================

                    OfertaGrupo.objects.filter(
                        oferta=oferta,
                    ).delete()

                    for grupo in grupos:

                        oferta_grupo = OfertaGrupo(
                            oferta=oferta,
                            grupo=grupo,
                            cupos=cupos,
                        )

                        oferta_grupo.full_clean()

                        oferta_grupo.save()

                    # ======================================
                    # 3. ACTUALIZAR HORARIO
                    # ======================================

                    horario.aula = aula
                    horario.dia = dia
                    horario.hora_inicio = hora_inicio
                    horario.hora_fin = hora_fin
                    horario.activo = True

                    # --------------------------------------
                    # Horario.clean() excluye su propio PK,
                    # por lo que no se detectará a sí mismo
                    # como conflicto.
                    # --------------------------------------

                    horario.full_clean()

                    horario.save()

            except ValidationError as error:

                agregar_errores_formulario(
                    form,
                    error,
                )

            else:

                if grupos.count() > 1:

                    mensaje = "Asignación transversal " "actualizada correctamente."

                else:

                    mensaje = "Asignación actualizada " "correctamente."

                messages.success(
                    request,
                    mensaje,
                )

                return redirect(volver_a)

    # ======================================================
    # GET
    # ======================================================

    else:

        form = NuevaAsignacionForm(
            programa=programa,
            semestre=semestre,
            centro_tutorial=centro,
            periodo=periodo,
            sede_inicial=sede_actual,
            initial={
                "asignatura": (oferta.asignatura_id),
                "profesor": (oferta.profesor_id),
                "modalidad": (oferta.modalidad_id),
                "grupos": [grupo.id for grupo in grupos_actuales],
                "cupos": (cupos_actuales),
                "sede": (sede_actual.id),
                "aula": (horario.aula_id),
                "dia": (horario.dia),
                "hora_inicio": (horario.hora_inicio),
                "hora_fin": (horario.hora_fin),
            },
        )

    # ======================================================
    # CONTEXTO
    # ======================================================

    context = {
        "form": form,
        "horario": horario,
        "oferta": oferta,
        "centro": centro,
        "programa": programa,
        "periodo": periodo,
        "semestre": semestre,
        "sede": sede_actual,
        "volver_a": volver_a,
    }

    return render(
        request,
        "horarios/editar_asignacion.html",
        context,
    )


# ==========================================================
# ELIMINAR ASIGNACIÓN
# ==========================================================


@login_required
@permission_required(
    "horarios.delete_horario",
    raise_exception=True,
)
def eliminar_asignacion(
    request,
    horario_id,
):

    horario = get_object_or_404(
        Horario.objects.select_related(
            "oferta",
            "oferta__asignatura",
            "oferta__profesor",
            "aula",
        ),
        pk=horario_id,
    )

    oferta = horario.oferta

    volver_a = request.POST.get("volver_a") or request.GET.get("volver_a") or "/"

    # ======================================================
    # INFORMACIÓN PARA LA CONFIRMACIÓN
    # ======================================================

    grupos = list(
        oferta.grupos.select_related(
            "programa",
        ).all()
    )

    # ======================================================
    # ELIMINAR
    # ======================================================

    if request.method == "POST":

        try:

            with transaction.atomic():

                # Guardamos información antes de eliminar.

                nombre_asignatura = oferta.asignatura.nombre

                # ------------------------------------------
                # Eliminamos el horario seleccionado.
                # ------------------------------------------

                horario.delete()

                # ------------------------------------------
                # ¿La oferta todavía tiene otros horarios?
                # ------------------------------------------

                tiene_otros_horarios = Horario.objects.filter(
                    oferta=oferta,
                ).exists()

                # ------------------------------------------
                # Si no quedan horarios, la Oferta Académica
                # ya no tiene razón de existir.
                #
                # Al eliminar OfertaAcademica se eliminarán
                # también sus OfertaGrupo mediante CASCADE.
                # ------------------------------------------

                if not tiene_otros_horarios:

                    oferta.delete()

        except Exception:

            messages.error(
                request,
                ("No fue posible eliminar la " "asignación."),
            )

        else:

            messages.success(
                request,
                (
                    f'La asignación "{nombre_asignatura}" '
                    "fue eliminada correctamente."
                ),
            )

        return redirect(volver_a)

    # ======================================================
    # CONFIRMACIÓN
    # ======================================================

    context = {
        "horario": horario,
        "oferta": oferta,
        "grupos": grupos,
        "volver_a": volver_a,
    }

    return render(
        request,
        "horarios/eliminar_asignacion.html",
        context,
    )


# ==========================================================
# EXPORTAR PROGRAMACIÓN A EXCEL
# ESTRUCTURA BASADA EN EL ARCHIVO REAL DE LA UNIVERSIDAD
# ==========================================================


@login_required
@permission_required(
    "horarios.view_horario",
    raise_exception=True,
)
def exportar_excel(request):
    centro_id = request.GET.get("centro")
    sede_id = request.GET.get("sede")
    facultad_id = request.GET.get("facultad")
    programa_id = request.GET.get("programa") or request.GET.get("carrera")
    periodo_id = request.GET.get("periodo")
    semestre_id = request.GET.get("semestre")
    asignatura_id = request.GET.get("asignatura")
    profesor_id = request.GET.get("profesor")
    modalidad_id = request.GET.get("modalidad")
    grupo_id = request.GET.get("grupo")
    aula_id = request.GET.get("aula")
    area_id = request.GET.get("area")
    dia = request.GET.get("dia")

    horarios_exportacion = Horario.objects.filter(
        activo=True,
    ).select_related(
        "aula",
        "aula__sede",
    )

    if sede_id:
        horarios_exportacion = horarios_exportacion.filter(aula__sede_id=sede_id)

    if aula_id:
        horarios_exportacion = horarios_exportacion.filter(aula_id=aula_id)

    if dia:
        horarios_exportacion = horarios_exportacion.filter(dia=dia)

    ofertas_grupo = OfertaGrupo.objects.filter(
        oferta__activo=True,
        grupo__activo=True,
        oferta__horarios__activo=True,
    )

    if centro_id:
        ofertas_grupo = ofertas_grupo.filter(grupo__centro_tutorial_id=centro_id)

    if facultad_id:
        ofertas_grupo = ofertas_grupo.filter(grupo__programa__facultad_id=facultad_id)

    if programa_id:
        ofertas_grupo = ofertas_grupo.filter(grupo__programa_id=programa_id)

    if periodo_id:
        ofertas_grupo = ofertas_grupo.filter(
            oferta__periodo_id=periodo_id,
            grupo__periodo_id=periodo_id,
        )

    if semestre_id:
        ofertas_grupo = ofertas_grupo.filter(grupo__semestre_id=semestre_id)

    if asignatura_id:
        ofertas_grupo = ofertas_grupo.filter(oferta__asignatura_id=asignatura_id)

    if profesor_id:
        ofertas_grupo = ofertas_grupo.filter(oferta__profesor_id=profesor_id)

    if modalidad_id:
        ofertas_grupo = ofertas_grupo.filter(oferta__modalidad_id=modalidad_id)

    if grupo_id:
        ofertas_grupo = ofertas_grupo.filter(grupo_id=grupo_id)

    if sede_id:
        ofertas_grupo = ofertas_grupo.filter(oferta__horarios__aula__sede_id=sede_id)

    if aula_id:
        ofertas_grupo = ofertas_grupo.filter(oferta__horarios__aula_id=aula_id)

    if dia:
        ofertas_grupo = ofertas_grupo.filter(oferta__horarios__dia=dia)

    ofertas_grupo = (
        ofertas_grupo.select_related(
            "oferta",
            "oferta__asignatura",
            "oferta__profesor",
            "oferta__periodo",
            "oferta__modalidad",
            "grupo",
            "grupo__centro_tutorial",
            "grupo__programa",
            "grupo__programa__facultad",
            "grupo__semestre",
        )
        .prefetch_related(
            Prefetch(
                "oferta__horarios",
                queryset=horarios_exportacion,
                to_attr="horarios_exportacion",
            )
        )
        .distinct()
        .order_by(
            "grupo__centro_tutorial__nombre",
            "grupo__programa__nombre",
            "grupo__semestre__numero",
            "oferta__asignatura__nombre",
            "grupo__codigo",
        )
    )

    planes = PlanEstudioAsignatura.objects.filter(
        activo=True,
    ).select_related(
        "area_academica",
    )

    if programa_id:
        planes = planes.filter(programa_id=programa_id)

    if semestre_id:
        planes = planes.filter(semestre_id=semestre_id)

    if asignatura_id:
        planes = planes.filter(asignatura_id=asignatura_id)

    mapa_planes = {
        (
            plan.programa_id,
            plan.asignatura_id,
            plan.semestre_id,
        ): plan
        for plan in planes
    }

    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "PROGRAMACIÓN"

    encabezados = [
        "Lugar de Desarrollo y/o Centro Tutorial",
        "Facultad",
        "Modalidad",
        "Programa Académico",
        "SEM",
        "Asignatura",
        "Grupo",
        "Cupos",
        "# Créditos",
        "Profesor",
        "Día",
        "Hora inicio",
        "Hora final",
        "Aula",
        "Sede",
        "Área Académica",
    ]

    for columna, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(
            row=1,
            column=columna,
            value=encabezado,
        )

        celda.font = Font(
            bold=True,
            color="FFFFFF",
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="212529",
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    fila_actual = 2

    for oferta_grupo in ofertas_grupo:
        oferta = oferta_grupo.oferta
        grupo = oferta_grupo.grupo

        plan = mapa_planes.get(
            (
                grupo.programa_id,
                oferta.asignatura_id,
                grupo.semestre_id,
            )
        )

        if area_id:
            if not plan or str(plan.area_academica_id) != str(area_id):
                continue

        for horario in oferta.horarios_exportacion:
            valores = [
                grupo.centro_tutorial.nombre,
                grupo.programa.facultad.nombre,
                oferta.modalidad.nombre,
                grupo.programa.nombre,
                grupo.semestre.get_numero_display(),
                oferta.asignatura.nombre,
                grupo.codigo,
                oferta_grupo.cupos,
                plan.creditos if plan else "",
                str(oferta.profesor),
                horario.get_dia_display(),
                horario.hora_inicio.strftime("%H:%M"),
                horario.hora_fin.strftime("%H:%M"),
                horario.aula.nombre,
                horario.aula.sede.nombre,
                (plan.area_academica.nombre if plan and plan.area_academica else ""),
            ]

            for columna, valor in enumerate(valores, start=1):
                hoja.cell(
                    row=fila_actual,
                    column=columna,
                    value=valor,
                )

            fila_actual += 1

    anchos = {
        "A": 35,
        "B": 25,
        "C": 18,
        "D": 32,
        "E": 10,
        "F": 38,
        "G": 18,
        "H": 10,
        "I": 12,
        "J": 30,
        "K": 15,
        "L": 14,
        "M": 14,
        "N": 18,
        "O": 25,
        "P": 25,
    }

    for columna, ancho in anchos.items():
        hoja.column_dimensions[columna].width = ancho

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:P{max(1, fila_actual - 1)}"

    nombre_archivo = "programacion_horarios.xlsx"

    response = HttpResponse(
        content_type=(
            "application/" "vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'

    workbook.save(response)

    return response
